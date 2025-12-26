"""
Image Augmentation and Feature Robustness Evaluation Module
=============================================================

This module provides tools for testing the robustness of feature detection
algorithms across various image augmentations.

Features:
- Image augmentation (rotation, scaling, blur, brightness, mirror)
- Feature matching across augmented versions
- Repeatability metrics
- Excel export of evaluation results

Author: AI Assistant
Date: November 2025
"""

import cv2
import numpy as np
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from skimage import transform as sktransform


@dataclass
class AugmentationParams:
    """Parameters for a specific augmentation."""
    name: str
    type: str  # 'rotation', 'scale', 'blur', 'brightness', 'mirror'
    value: float  # parameter value (degrees, scale factor, sigma, percentage, flip code)


@dataclass
class EvaluationResult:
    """Results from feature robustness evaluation."""
    augmentation: str
    feature_count_original: int
    feature_count_augmented: int
    match_count: int
    match_ratio: float
    repeatability: float
    execution_time_ms: float


class ImageAugmenter:
    """
    Provides various image augmentation operations.
    """
    
    @staticmethod
    def rotate(image: np.ndarray, degrees: float) -> np.ndarray:
        """
        Rotate image by specified degrees.
        
        Args:
            image: Input image (BGR or grayscale)
            degrees: Rotation angle in degrees
            
        Returns:
            Rotated image
        """
        h, w = image.shape[:2]
        center = (w // 2, h // 2)
        
        # Calculate rotation matrix
        matrix = cv2.getRotationMatrix2D(center, degrees, 1.0)
        
        # Calculate new image size to avoid cropping
        cos = np.abs(matrix[0, 0])
        sin = np.abs(matrix[0, 1])
        new_w = int(h * sin + w * cos)
        new_h = int(h * cos + w * sin)
        
        # Adjust rotation matrix for new size
        matrix[0, 2] += (new_w / 2) - center[0]
        matrix[1, 2] += (new_h / 2) - center[1]
        
        # Perform rotation
        rotated = cv2.warpAffine(image, matrix, (new_w, new_h), 
                                 borderMode=cv2.BORDER_CONSTANT, 
                                 borderValue=(0, 0, 0))
        return rotated
    
    @staticmethod
    def scale(image: np.ndarray, factor: float) -> np.ndarray:
        """
        Scale image by specified factor.
        
        Args:
            image: Input image
            factor: Scale factor (e.g., 0.5 = half size, 2.0 = double size)
            
        Returns:
            Scaled image
        """
        h, w = image.shape[:2]
        new_w = int(w * factor)
        new_h = int(h * factor)
        scaled = cv2.resize(image, (new_w, new_h), interpolation=cv2.INTER_LINEAR)
        return scaled
    
    @staticmethod
    def blur(image: np.ndarray, sigma: float = 2.0) -> np.ndarray:
        """
        Apply Gaussian blur.
        
        Args:
            image: Input image
            sigma: Standard deviation for Gaussian kernel
            
        Returns:
            Blurred image
        """
        # Calculate kernel size from sigma (ensure odd number)
        ksize = int(6 * sigma + 1)
        if ksize % 2 == 0:
            ksize += 1
        
        blurred = cv2.GaussianBlur(image, (ksize, ksize), sigma)
        return blurred
    
    @staticmethod
    def adjust_brightness(image: np.ndarray, percentage: float) -> np.ndarray:
        """
        Adjust image brightness.
        
        Args:
            image: Input image
            percentage: Brightness adjustment (-100 to 100)
                       Negative = darker, Positive = brighter
            
        Returns:
            Brightness-adjusted image
        """
        factor = 1.0 + (percentage / 100.0)
        adjusted = cv2.convertScaleAbs(image, alpha=factor, beta=0)
        return adjusted
    
    @staticmethod
    def mirror(image: np.ndarray, horizontal: bool = True, vertical: bool = False) -> np.ndarray:
        """
        Mirror (flip) image.
        
        Args:
            image: Input image
            horizontal: Flip horizontally
            vertical: Flip vertically
            
        Returns:
            Mirrored image
        """
        if horizontal and vertical:
            flip_code = -1
        elif horizontal:
            flip_code = 1
        elif vertical:
            flip_code = 0
        else:
            return image.copy()
        
        mirrored = cv2.flip(image, flip_code)
        return mirrored
    
    @staticmethod
    def noise(image: np.ndarray, std: float = 10.0) -> np.ndarray:
        """
        Add Gaussian noise to image.
        
        Args:
            image: Input image
            std: Standard deviation of noise
            
        Returns:
            Noisy image
        """
        noise = np.random.normal(0, std, image.shape).astype(np.float32)
        noisy = image.astype(np.float32) + noise
        noisy = np.clip(noisy, 0, 255).astype(np.uint8)
        return noisy


class FeatureRobustnessEvaluator:
    """
    Evaluates the robustness of feature detection across augmentations.
    """
    
    def __init__(self, method: str = "ORB"):
        """
        Initialize evaluator.
        
        Args:
            method: Feature detection method ('ORB', 'SIFT', 'AKAZE')
        """
        self.method = method
        self.detector = None
        self.matcher = None
        self._setup_detector()
    
    def _setup_detector(self):
        """Setup feature detector and matcher based on method."""
        if self.method == "ORB":
            self.detector = cv2.ORB_create(nfeatures=500)
            self.matcher = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=False)
        elif self.method == "SIFT":
            if hasattr(cv2, 'SIFT_create'):
                self.detector = cv2.SIFT_create(nfeatures=500)
                self.matcher = cv2.BFMatcher(cv2.NORM_L2, crossCheck=False)
            else:
                # Fallback to ORB
                print("SIFT not available, using ORB")
                self.method = "ORB"
                self._setup_detector()
                return
        elif self.method == "AKAZE":
            self.detector = cv2.AKAZE_create()
            self.matcher = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=False)
        else:
            raise ValueError(f"Unknown method: {self.method}")
    
    def detect_and_compute(self, image: np.ndarray) -> Tuple[List, np.ndarray]:
        """
        Detect keypoints and compute descriptors.
        
        Args:
            image: Input image (BGR or grayscale)
            
        Returns:
            Tuple of (keypoints, descriptors)
        """
        # Convert to grayscale if needed
        if len(image.shape) == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        else:
            gray = image
        
        # Detect and compute
        keypoints, descriptors = self.detector.detectAndCompute(gray, None)
        
        return keypoints, descriptors
    
    def match_features(self, desc1: np.ndarray, desc2: np.ndarray, 
                      ratio_threshold: float = 0.75) -> Tuple[int, List]:
        """
        Match features using ratio test.
        
        Args:
            desc1: Descriptors from first image
            desc2: Descriptors from second image
            ratio_threshold: Lowe's ratio test threshold
            
        Returns:
            Tuple of (good_match_count, good_matches)
        """
        if desc1 is None or desc2 is None or len(desc1) == 0 or len(desc2) == 0:
            return 0, []
        
        # Find k=2 nearest matches
        matches = self.matcher.knnMatch(desc1, desc2, k=2)
        
        # Apply Lowe's ratio test
        good_matches = []
        for match_pair in matches:
            if len(match_pair) == 2:
                m, n = match_pair
                if m.distance < ratio_threshold * n.distance:
                    good_matches.append(m)
        
        return len(good_matches), good_matches
    
    def compute_repeatability(self, kp1: List, kp2: List,  
                             transform_matrix: Optional[np.ndarray] = None,
                             epsilon: float = 5.0) -> float:
        """
        Compute keypoint repeatability (percentage of keypoints that reappear).
        
        Args:
            kp1: Keypoints from original image
            kp2: Keypoints from augmented image
            transform_matrix: Transformation matrix (for geometric verification)
            epsilon: Distance threshold in pixels
            
        Returns:
            Repeatability score (0 to 1)
        """
        if len(kp1) == 0 or len(kp2) == 0:
            return 0.0
        
        # Convert keypoints to numpy arrays
        pts1 = np.float32([kp.pt for kp in kp1])
        pts2 = np.float32([kp.pt for kp in kp2])
        
        # If transform provided, apply it to points
        if transform_matrix is not None:
            pts1_transformed = cv2.transform(np.array([pts1]), transform_matrix)[0]
        else:
            pts1_transformed = pts1
        
        # Count matches within epsilon
        matched = 0
        for pt1 in pts1_transformed:
            # Calculate distances to all points in kp2
            dists = np.linalg.norm(pts2 - pt1, axis=1)
            if np.min(dists) <= epsilon:
                matched += 1
        
        repeatability = matched / len(kp1) if len(kp1) > 0 else 0.0
        return repeatability
    
    def evaluate_augmentation(self, original: np.ndarray, augmented: np.ndarray,
                             augmentation_name: str) -> EvaluationResult:
        """
        Evaluate feature detection robustness for one augmentation.
        
        Args:
            original: Original image
            augmented: Augmented image
            augmentation_name: Name of augmentation
            
        Returns:
            EvaluationResult with metrics
        """
        import time
        
        # Detect features in original
        start_time = time.time()
        kp1, desc1 = self.detect_and_compute(original)
        exec_time_orig = (time.time() - start_time) * 1000
        
        # Detect features in augmented
        start_time = time.time()
        kp2, desc2 = self.detect_and_compute(augmented)
        exec_time_aug = (time.time() - start_time) * 1000
        
        # Match features
        match_count, _ = self.match_features(desc1, desc2)
        
        # Calculate match ratio
        total_features = max(len(kp1), len(kp2))
        match_ratio = match_count / total_features if total_features > 0 else 0.0
        
        # Calculate repeatability
        repeatability = self.compute_repeatability(kp1, kp2)
        
        # Average execution time
        avg_exec_time = (exec_time_orig + exec_time_aug) / 2
        
        return EvaluationResult(
            augmentation=augmentation_name,
            feature_count_original=len(kp1),
            feature_count_augmented=len(kp2),
            match_count=match_count,
            match_ratio=match_ratio,
            repeatability=repeatability,
            execution_time_ms=avg_exec_time
        )
    
    def evaluate_all_augmentations(self, image: np.ndarray, 
                                   augmentations: List[AugmentationParams]) -> List[EvaluationResult]:
        """
        Evaluate feature robustness across multiple augmentations.
        
        Args:
            image: Original image
            augmentations: List of augmentation parameters
            
        Returns:
            List of evaluation results
        """
        results = []
        augmenter = ImageAugmenter()
        
        for aug_params in augmentations:
            # Apply augmentation
            if aug_params.type == 'rotation':
                augmented = augmenter.rotate(image, aug_params.value)
            elif aug_params.type == 'scale':
                augmented = augmenter.scale(image, aug_params.value)
            elif aug_params.type == 'blur':
                augmented = augmenter.blur(image, aug_params.value)
            elif aug_params.type == 'brightness':
                augmented = augmenter.adjust_brightness(image, aug_params.value)
            elif aug_params.type == 'mirror_h':
                augmented = augmenter.mirror(image, horizontal=True, vertical=False)
            elif aug_params.type == 'mirror_v':
                augmented = augmenter.mirror(image, horizontal=False, vertical=True)
            elif aug_params.type == 'noise':
                augmented = augmenter.noise(image, aug_params.value)
            else:
                print(f"Unknown augmentation type: {aug_params.type}")
                continue
            
            # Evaluate
            result = self.evaluate_augmentation(image, augmented, aug_params.name)
            results.append(result)
        
        return results


class ExcelExporter:
    """
    Exports evaluation results to Excel with formatting and charts.
    """
    
    @staticmethod
    def export_results(results: List[EvaluationResult], 
                      filepath: str,
                      method_name: str = "ORB",
                      image_name: str = "Unknown"):
        """
        Export evaluation results to Excel file.
        
        Args:
            results: List of evaluation results
            filepath: Output Excel file path
            method_name: Feature detection method name
            image_name: Name of the tested image
        """
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Feature Robustness"
        
        # Header
        ws['A1'] = f"Feature Robustness Evaluation - {method_name}"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells('A1:H1')
        
        # Metadata
        ws['A2'] = f"Image: {image_name}"
        ws['A3'] = f"Method: {method_name}"
        ws['A4'] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A5'] = f"Total Augmentations: {len(results)}"
        
        # Column headers
        headers = [
            "Augmentation",
            "Original Features",
            "Augmented Features",
            "Matches",
            "Match Ratio (%)",
            "Repeatability (%)",
            "Exec Time (ms)",
            "Status"
        ]
        
        header_row = 7
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for row_idx, result in enumerate(results, start=header_row + 1):
            ws.cell(row=row_idx, column=1, value=result.augmentation)
            ws.cell(row=row_idx, column=2, value=result.feature_count_original)
            ws.cell(row=row_idx, column=3, value=result.feature_count_augmented)
            ws.cell(row=row_idx, column=4, value=result.match_count)
            ws.cell(row=row_idx, column=5, value=round(result.match_ratio * 100, 2))
            ws.cell(row=row_idx, column=6, value=round(result.repeatability * 100, 2))
            ws.cell(row=row_idx, column=7, value=round(result.execution_time_ms, 2))
            
            # Status based on match ratio
            if result.match_ratio > 0.7:
                status = "Excellent"
                color = "92D050"
            elif result.match_ratio > 0.5:
                status = "Good"
                color = "FFD966"
            elif result.match_ratio > 0.3:
                status = "Fair"
                color = "FFC000"
            else:
                status = "Poor"
                color = "FF6B6B"
            
            cell = ws.cell(row=row_idx, column=8, value=status)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(bold=True)
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary statistics
        summary_row = len(results) + header_row + 2
        ws.cell(row=summary_row, column=1, value="Summary Statistics").font = Font(bold=True)
        
        if results:
            avg_match_ratio = np.mean([r.match_ratio for r in results]) * 100
            avg_repeatability = np.mean([r.repeatability for r in results]) * 100
            avg_exec_time = np.mean([r.execution_time_ms for r in results])
            
            ws.cell(row=summary_row + 1, column=1, value="Average Match Ratio:")
            ws.cell(row=summary_row + 1, column=2, value=f"{avg_match_ratio:.2f}%")
            
            ws.cell(row=summary_row + 2, column=1, value="Average Repeatability:")
            ws.cell(row=summary_row + 2, column=2, value=f"{avg_repeatability:.2f}%")
            
            ws.cell(row=summary_row + 3, column=1, value="Average Execution Time:")
            ws.cell(row=summary_row + 3, column=2, value=f"{avg_exec_time:.2f} ms")
        
        # Add chart
        if len(results) > 0:
            chart = BarChart()
            chart.title = "Feature Matching Performance"
            chart.y_axis.title = "Percentage (%)"
            chart.x_axis.title = "Augmentation"
            
            # Data for chart (Match Ratio and Repeatability)
            data = Reference(ws, min_col=5, min_row=header_row, max_row=len(results) + header_row, max_col=6)
            cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=len(results) + header_row)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 20
            
            ws.add_chart(chart, f"A{summary_row + 6}")
        
        # Save workbook
        wb.save(filepath)
        print(f"Excel report saved to: {filepath}")


# Predefined augmentation sets
STANDARD_AUGMENTATIONS = [
    AugmentationParams("Rotate 15°", "rotation", 15),
    AugmentationParams("Rotate 30°", "rotation", 30),
    AugmentationParams("Rotate 45°", "rotation", 45),
    AugmentationParams("Rotate 90°", "rotation", 90),
    AugmentationParams("Scale 0.7x", "scale", 0.7),
    AugmentationParams("Scale 1.3x", "scale", 1.3),
    AugmentationParams("Blur σ=2", "blur", 2.0),
    AugmentationParams("Blur σ=4", "blur", 4.0),
    AugmentationParams("Brightness +30%", "brightness", 30),
    AugmentationParams("Brightness -30%", "brightness", -30),
    AugmentationParams("Mirror Horizontal", "mirror_h", 1),
    AugmentationParams("Mirror Vertical", "mirror_v", 1),
    AugmentationParams("Noise σ=10", "noise", 10),
]

